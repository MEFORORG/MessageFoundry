# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 MessageFoundry Organization and contributors
"""Render acceptance results — console summary, Markdown, CSV, and optional spreadsheet write-back.

The Markdown/CSV writers are dependency-free. The spreadsheet write-back (filling the matrix
workbook's Status column) needs ``openpyxl``; it is imported lazily so the rest of the harness — and
the project's own test run — never depends on it.
"""

from __future__ import annotations

import csv
import io
from collections import Counter
from pathlib import Path
from typing import Sequence

from harness.acceptance.matrix import SECTIONS, Status
from harness.acceptance.runner import RowResult

#: A run is a failure if any executed row failed or errored; MANUAL/SKIP never fail the run.
_FAILING = (Status.FAIL, Status.ERROR)


def summarize(results: Sequence[RowResult]) -> Counter[Status]:
    """Count results by status."""
    return Counter(r.status for r in results)


def exit_code(results: Sequence[RowResult]) -> int:
    """0 unless a row FAILed or ERRORed."""
    return 1 if any(r.status in _FAILING for r in results) else 0


def render_console(results: Sequence[RowResult]) -> str:
    """A compact, aligned one-line-per-row summary for stdout."""
    counts = summarize(results)
    lines = [
        f"{r.status.value:<6} {r.row.id:<4} {r.row.title}" + (f"  [{r.detail}]" if r.detail else "")
        for r in results
    ]
    tally = "  ".join(f"{s.value}={counts.get(s, 0)}" for s in Status)
    lines.append("")
    lines.append(f"  {tally}   (exit {exit_code(results)})")
    return "\n".join(lines)


def render_markdown(results: Sequence[RowResult]) -> str:
    """A sectioned Markdown report mirroring the matrix layout."""
    counts = summarize(results)
    out: list[str] = ["# WIN2025 Acceptance — Results", ""]
    out.append(
        "Tally: "
        + " · ".join(f"**{s.value}** {counts.get(s, 0)}" for s in Status)
        + f" · exit `{exit_code(results)}`"
    )
    out.append("")
    by_section: dict[str, list[RowResult]] = {letter: [] for letter in SECTIONS}
    for r in results:
        by_section.setdefault(r.row.section, []).append(r)
    for letter, rows in by_section.items():
        if not rows:
            continue
        out.append(f"## {letter}. {SECTIONS.get(letter, '')}")
        out.append("")
        out.append("| ID | Test | Per-DB? | Status | Detail |")
        out.append("|---|---|---|---|---|")
        for r in rows:
            detail = (r.detail or "").replace("|", "\\|")
            out.append(
                f"| {r.row.id} | {r.row.title.replace('|', chr(92) + '|')} "
                f"| {r.row.per_db} | {r.status.value} | {detail} |"
            )
        out.append("")
    return "\n".join(out)


def render_csv(results: Sequence[RowResult]) -> str:
    """A flat CSV row per result."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(
        ["id", "section", "title", "per_db", "coverage", "status", "detail", "evidence"]
    )
    for r in results:
        writer.writerow(
            [
                r.row.id,
                r.row.section,
                r.row.title,
                r.row.per_db,
                r.row.coverage.value,
                r.status.value,
                r.detail,
                r.evidence,
            ]
        )
    return buf.getvalue()


def write_xlsx_status(results: Sequence[RowResult], xlsx_path: Path) -> int:
    """Write each result's Status (+ detail) back into the matrix workbook, matched by row ID.

    Returns the number of rows updated. Raises ``RuntimeError`` if openpyxl is unavailable or the
    workbook has no recognisable ``ID`` / ``Status`` header.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # not a project dependency — best-effort only
        raise RuntimeError(
            "openpyxl is required for spreadsheet write-back; run with --report-md/--report-csv instead"
        ) from exc

    wb = load_workbook(xlsx_path)
    ws = wb.active
    by_id = {r.row.id: r for r in results}

    # Locate the header row + the ID / Status / result columns.
    id_col = status_col = result_col = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        labels = {str(c.value).strip().lower(): c.column for c in row if c.value is not None}
        if "id" in labels and "status" in labels:
            header_row = row[0].row
            id_col = labels["id"]
            status_col = labels["status"]
            result_col = next(
                (col for label, col in labels.items() if "result" in label or "evidence" in label),
                None,
            )
            break
    if header_row is None or id_col is None or status_col is None:
        raise RuntimeError(f"no 'ID'/'Status' header found in {xlsx_path.name}")

    updated = 0
    for row in ws.iter_rows(min_row=header_row + 1):
        cell = row[id_col - 1]
        result = by_id.get(str(cell.value).strip()) if cell.value is not None else None
        if result is None:
            continue
        ws.cell(row=cell.row, column=status_col, value=result.status.value)
        if result_col is not None:
            ws.cell(row=cell.row, column=result_col, value=result.detail)
        updated += 1

    wb.save(xlsx_path)
    return updated
