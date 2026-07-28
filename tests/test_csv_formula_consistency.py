# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 MessageFoundry Organization and contributors
"""ASVS 1.2.10 — one spreadsheet formula-injection rule, enforced across every writer in the tree.

The finding this closes was not "the escaping is missing" but "the escaping **diverges**": the engine
audit export triggered on ``= + - @ TAB CR LF`` (**no NUL**) while four copies triggered on
``= + - @ TAB CR NUL`` (**no LF**) and tested only ``value[:1]``, so ``"   =evil()"`` rode through
them. The attacker-influenced writer — the audit CSV, whose actor/action/client/detail come from
authenticated user input and request headers — was the one missing NUL.

Copy-paste is what let the rule drift, so this module is the gate that a copy cannot pass:

* every writer's trigger set is *the* canonical set, and
* every writer produces **identical output** over one shared vector list, and
* every module in the repo that constructs a ``csv`` writer — through **any** binding of the module
  (``import csv``, ``import csv as _csv``, ``from csv import writer``) — or imports a spreadsheet
  library (``openpyxl``/``xlsxwriter``/…) is a recorded one, so a **sixth** writer cannot appear
  outside the gate. (What that matcher does *not* prove: it reads static import statements, so a
  dynamically-imported writer — ``importlib.import_module("csv")`` — or a spreadsheet library outside
  :data:`_SPREADSHEET_LIBS` is invisible to it, as it is to any static scan.) And
* the requirement's "or other spreadsheet formats (such as XLS, XLSX, or ODF)" half is covered: the
  acceptance harness's ``.xlsx`` write-back must not store a detail as a live formula — asserted by a
  test that **stubs** openpyxl, because openpyxl is in no extra, no lock file and no CI workflow, so
  an ``importorskip`` test for it never executes anywhere.

The engine and the harness keep separate modules on purpose (a test harness must not import engine
internals, CLAUDE.md §4/§10); the tests below are what hold the two mirrors together.
"""

from __future__ import annotations

import ast
import csv
import io
import sys
import types
from collections.abc import Callable
from pathlib import Path

import pytest

import harness._spreadsheet as harness_rule
import messagefoundry.spreadsheet as engine_rule
from harness._spreadsheet import xlsx_cell_text
from harness.acceptance import report as acceptance_report
from harness.acceptance.matrix import MATRIX, Status
from harness.acceptance.report import render_csv, write_xlsx_status
from harness.acceptance.runner import RowResult
from harness.load import report as load_report
from harness.load.connscale import report as connscale_report
from messagefoundry.api import auth_routes
from messagefoundry.config import codeset_edit

_REPO = Path(__file__).resolve().parent.parent

#: The canonical rule: ``= + - @`` (a spreadsheet evaluates them) plus TAB / CR / LF / NUL (an
#: importer can strip one and expose the trigger behind it).
_CANONICAL = frozenset("=+-@\t\r\n\x00")

#: The five spreadsheet writers, each as ``(label, trigger set, escape helper)``. Named through the
#: writer modules' own attributes — an alias swapped back to a local copy fails the identity tests.
_WRITERS: list[tuple[str, frozenset[str], Callable[[object], object]]] = [
    ("api/auth_routes (audit export)", auth_routes._CSV_FORMULA_TRIGGERS, auth_routes._csv_safe),
    ("config/codeset_edit", codeset_edit._CSV_FORMULA_TRIGGERS, codeset_edit._spreadsheet_safe),
    (
        "harness/acceptance/report",
        acceptance_report._CSV_FORMULA_TRIGGERS,
        acceptance_report._spreadsheet_safe,
    ),
    ("harness/load/report", load_report._CSV_FORMULA_TRIGGERS, load_report._spreadsheet_safe),
    (
        "harness/load/connscale/report",
        connscale_report._CSV_FORMULA_TRIGGERS,
        connscale_report._spreadsheet_safe,
    ),
]

#: One shared vector list, run through every writer. The first block is exactly the set that used to
#: diverge; the second is the importer-strips-it class the raw-first-character rule alone misses; the
#: third must pass through untouched (over-escaping a benign cell is its own defect).
_HOSTILE = [
    "=cmd|'/c calc'!A1",
    "+1+1",
    "-2+3",
    "@SUM(A1)",
    "\t=evil()",
    "\r=evil()",
    "\n=evil()",  # LF: escaped only by the engine writer before this change
    "\x00=evil()",  # NUL: escaped only by the four copies before this change
    "\x00",
    "   =evil()",  # trigger behind spaces: escaped only by the engine writer before this change
    "\t   =evil()",
    "=",
    " =evil()",  # NBSP (Unicode whitespace)
    "​=evil()",  # zero-width space — not whitespace, so str.lstrip does not reach the "="
    "﻿=evil()",  # BOM / ZWNBSP
    "\x01=evil()",  # SOH: a C0 control, in the same importer-drops-it class as NUL
]
_BENIGN = ["alice", "message_view", '{"n": 1}', "", " ", "12345", "PASS", "n/a", "a=b", "x - y"]
#: Non-strings a writer may pipe through (the audit export writes a float ts; the load reports write
#: ints/floats): they must pass through, not raise — which is why the rule takes ``object``.
_NON_STR: list[object] = [100.0, 7, True, None]


def test_every_writer_uses_the_canonical_trigger_set() -> None:
    for label, triggers, _helper in _WRITERS:
        assert set(triggers) == set(_CANONICAL), label


def test_every_writer_produces_identical_output_over_the_shared_vectors() -> None:
    # The real gate: not "each is non-empty" but "all five agree, cell for cell".
    for vector in [*_HOSTILE, *_BENIGN, *_NON_STR]:
        outputs = {label: helper(vector) for label, _t, helper in _WRITERS}
        distinct = set(outputs.values())
        assert len(distinct) == 1, f"writers disagree on {vector!r}: {outputs}"


def test_hostile_vectors_are_actually_neutralized() -> None:
    # Anti-vacuity for the identity test above — five writers could agree by all being broken.
    for vector in _HOSTILE:
        for label, _t, helper in _WRITERS:
            assert helper(vector) == "'" + vector, f"{label} did not neutralize {vector!r}"


def test_benign_and_non_string_cells_pass_through_untouched() -> None:
    for vector in [*_BENIGN, *_NON_STR]:
        for label, _t, helper in _WRITERS:
            assert helper(vector) == vector, f"{label} over-escaped {vector!r}"


def test_neutralization_is_idempotent() -> None:
    # A report regenerated from an already-escaped value must not grow a second apostrophe.
    for vector in _HOSTILE:
        for _label, _t, helper in _WRITERS:
            once = helper(vector)
            assert helper(once) == once


def test_engine_and_harness_mirrors_agree() -> None:
    # The two modules are separate by design (CLAUDE.md §4/§10 keeps the harness off engine
    # internals); this is what keeps the copy honest.
    assert engine_rule.SPREADSHEET_FORMULA_TRIGGERS == harness_rule.SPREADSHEET_FORMULA_TRIGGERS
    assert engine_rule._LEADING_NOISE == harness_rule._LEADING_NOISE
    for vector in [*_HOSTILE, *_BENIGN, *_NON_STR]:
        assert engine_rule.spreadsheet_safe(vector) == harness_rule.spreadsheet_safe(vector)
    for text in [*_HOSTILE, *_BENIGN]:
        assert engine_rule.spreadsheet_lead(text) == harness_rule.spreadsheet_lead(text)


# --- RFC 4180 §2.6/2.7 quoting: asserted, not changed ------------------------


@pytest.mark.parametrize("dialect_kwargs", [{}, {"lineterminator": "\n"}], ids=["default", "lf"])
def test_stdlib_quoting_is_uniform_across_the_writers_dialects(dialect_kwargs: dict) -> None:
    # Both dialects in use (harness/acceptance/report.py passes lineterminator="\n"; the other four
    # take the default) already quote an embedded comma / quote / CR / LF and round-trip it intact, so
    # neutralization is the only thing this cell has to add. Asserted so a dialect change is visible.
    for cell in ["a,b", 'a"b', "a\nb", "a\rb", "a\r\nb"]:
        buf = io.StringIO(newline="")
        csv.writer(buf, **dialect_kwargs).writerow([cell, "z"])
        rendered = buf.getvalue()
        assert rendered.startswith('"'), (cell, rendered)
        assert next(csv.reader(io.StringIO(rendered, newline="")))[0] == cell


# --- the sixth-writer gate ---------------------------------------------------

#: Every spreadsheet writer in the repo — ``csv`` writers **and** spreadsheet-library importers — with
#: why it is safe. A new writer must be added here deliberately, which is the moment to route its
#: free-text cells through the rule.
_RECORDED_SPREADSHEET_WRITERS = {
    "messagefoundry/api/auth_routes.py": "escapes every cell via _csv_safe (attacker-influenced)",
    "messagefoundry/config/codeset_edit.py": "escapes the header row and every cell",
    "harness/acceptance/report.py": (
        "CSV: escapes title/detail/evidence; other columns are enum/ids. XLSX: the detail goes "
        "through xlsx_cell_text and the written cell is pinned to data_type 's'"
    ),
    "harness/load/report.py": "escapes profile/phase/kind; other columns are numeric",
    "harness/load/connscale/report.py": "escapes profile/claim_mode/sweep_mode; others numeric",
    "scripts/security/vuln_metrics.py": (
        "no free-text cell: every value is an ISO date, a formatted number, a count, or an 'n/a: …' "
        "string, so no cell can begin with '=', '+', '@', TAB/CR/LF or NUL. A formatted median CAN "
        "begin with '-' (a clock-skewed createdAt->mergedAt delta renders as e.g. -0.1) — a benign "
        "negative number in a metrics file, not a DDE payload. It is also stdlib-only by design (the "
        "vuln-metrics workflow runs it with no install), so it cannot import the engine rule. Route "
        "it through spreadsheet_safe if a free-text column, or the engine on sys.path, ever lands."
    ),
}

#: Libraries that write a *binary* spreadsheet format. The ASVS text covers "CSV **or other
#: spreadsheet formats (such as XLS, XLSX, or ODF)**", so a module importing one of these is a
#: spreadsheet writer for this gate's purposes even though it constructs no ``csv`` writer.
_SPREADSHEET_LIBS = frozenset({"openpyxl", "xlsxwriter", "xlwt", "odf", "odfpy", "pyexcel"})

#: `.claude` matters and is not cosmetic: the documented parallel-session workflow (docs/WORKTREES.md)
#: places sibling worktrees at `.claude/worktrees/<name>/`, i.e. NESTED INSIDE this checkout. This gate
#: walks `_REPO.rglob("*.py")`, so with a sibling session active it scanned a SECOND FULL COPY of the
#: repo — 5,712 extra .py files, at a DIFFERENT COMMIT — and failed on writers registered over there.
#: CI never sees this (no nested worktrees), so it reds only the local run, which is where people
#: iterate; the natural response is to learn to ignore this test, and then it guards nothing.
_SKIP_DIRS = {
    ".venv",
    ".git",
    ".claude",
    "node_modules",
    "__pycache__",
    ".mypy_cache",
    ".pytest_cache",
}


def _spreadsheet_writer_sites_in(source: str) -> bool:
    """Whether ``source`` constructs a ``csv`` writer or imports a spreadsheet library (AST, not grep).

    The ``csv`` binding is resolved **per module** first. Matching only ``ast.Attribute`` on the
    literal name ``csv`` — as this gate first did — walks straight past ``import csv as _csv`` (a
    pattern that already exists in this tree) and ``from csv import writer, DictWriter``, so a new
    writer could land in either form with the gate green.
    """
    tree = ast.parse(source)
    csv_modules: set[str] = set()  # local names bound to the csv MODULE
    csv_writers: set[str] = set()  # local names bound to csv.writer / csv.DictWriter directly
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                root = alias.name.split(".")[0]
                if alias.name == "csv":
                    csv_modules.add(alias.asname or "csv")
                if root in _SPREADSHEET_LIBS:
                    return True
        elif isinstance(node, ast.ImportFrom) and node.level == 0 and node.module:
            if node.module == "csv":
                csv_writers |= {
                    alias.asname or alias.name
                    for alias in node.names
                    if alias.name in {"writer", "DictWriter"}
                }
            if node.module.split(".")[0] in _SPREADSHEET_LIBS:
                return True
    for node in ast.walk(tree):
        if not isinstance(node, ast.Call):
            continue
        fn = node.func
        if (
            isinstance(fn, ast.Attribute)
            and fn.attr in {"writer", "DictWriter"}
            and isinstance(fn.value, ast.Name)
            and fn.value.id in csv_modules
        ):
            return True
        if isinstance(fn, ast.Name) and fn.id in csv_writers:
            return True
    return False


def _spreadsheet_writer_sites() -> set[str]:
    """Repo-relative paths of every module that writes a spreadsheet format."""
    found: set[str] = set()
    for path in _REPO.rglob("*.py"):
        if _SKIP_DIRS & set(path.parts):
            continue
        try:
            source = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:  # pragma: no cover - not expected in-tree
            continue
        try:
            if _spreadsheet_writer_sites_in(source):
                found.add(path.relative_to(_REPO).as_posix())
        except SyntaxError:  # pragma: no cover - not expected in-tree
            continue
    return found


def test_no_unrecorded_spreadsheet_writer_exists() -> None:
    sites = _spreadsheet_writer_sites() - {Path(__file__).relative_to(_REPO).as_posix()}
    unrecorded = sorted(sites - _RECORDED_SPREADSHEET_WRITERS.keys())
    assert not unrecorded, (
        "new spreadsheet writer(s) outside the ASVS 1.2.10 gate — route free-text cells through the "
        f"canonical rule and record them here: {unrecorded}"
    )
    stale = sorted(_RECORDED_SPREADSHEET_WRITERS.keys() - sites)
    assert not stale, f"recorded spreadsheet writer(s) that no longer exist: {stale}"


def test_the_writer_gate_sees_every_binding_of_the_csv_module() -> None:
    # Anti-vacuity for the gate above, one vector per bypass the attribute-only matcher had. The
    # aliased-import form is not hypothetical: it is already used in tests/test_acceptance_framework.py.
    for source in (
        "import csv\ncsv.writer(f)\n",
        "import csv as _csv\n_csv.writer(f)\n",
        "import csv as _csv\n_csv.DictWriter(f, fieldnames=[])\n",
        "from csv import writer\nwriter(f)\n",
        "from csv import DictWriter as DW\nDW(f, fieldnames=[])\n",
        "import openpyxl\nopenpyxl.Workbook()\n",
        "from openpyxl import load_workbook\nload_workbook(p)\n",
        "def f():\n    import xlsxwriter\n    return xlsxwriter\n",
    ):
        assert _spreadsheet_writer_sites_in(source), source
    # ...and does not fire on a csv READER, a same-named attribute on something else, or a shadow.
    for source in (
        "import csv\ncsv.reader(f)\n",
        "import shapefile\nshapefile.writer(f)\n",
        "from mymod import writer\nwriter(f)\n",
    ):
        assert not _spreadsheet_writer_sites_in(source), source


# --- the XLSX half ("or other spreadsheet formats") --------------------------


def test_xlsx_cell_text_neutralizes_a_leading_equals() -> None:
    # openpyxl stores any '='-leading string as a FORMULA cell, so an '='-leading detail executed on
    # open. Dependency-free by design: openpyxl appears in no extra, no lock file and no CI workflow —
    # it is installed by hand only on the WIN2025 acceptance box (docs/testing/WIN2025-ACCEPTANCE.md),
    # so an `importorskip("openpyxl")` test never executes ANYWHERE, in CI or locally. The regression
    # for this defect therefore must not need the library.
    assert xlsx_cell_text("=cmd|'/c calc'!A1") == "'=cmd|'/c calc'!A1"
    assert xlsx_cell_text("=") == "'="
    assert not xlsx_cell_text("=1+1").startswith("=")


def test_xlsx_cell_text_leaves_the_csv_only_triggers_alone() -> None:
    # The XLSX threat scope is genuinely narrower than CSV's: openpyxl stores '+'/'-'/'@'/TAB-leading
    # strings as inline strings and Excel does not re-parse them the way it re-parses a CSV import.
    # Escaping them here would mutate report text that was never dangerous.
    for text in ["+1+1", "-2+3", "@SUM(A1)", "\t=evil()", "   =evil()", "alice", ""]:
        assert xlsx_cell_text(text) == text


def test_xlsx_cell_text_drops_characters_openpyxl_refuses_to_store() -> None:
    # A NUL (or any of \x01-\x08, \x0b, \x0c, \x0e-\x1f) in a detail made openpyxl raise
    # IllegalCharacterError, which derives straight from Exception and so escaped the acceptance CLI's
    # (RuntimeError, OSError) handler — killing the run with a traceback instead of "xlsx write-back
    # failed" + exit 2. TAB/LF/CR are legal and preserved.
    assert xlsx_cell_text("\x00=evil()") == "'=evil()"
    assert xlsx_cell_text("a\x01b\x1fc") == "abc"
    assert xlsx_cell_text("a\tb\nc\rd") == "a\tb\nc\rd"


def test_acceptance_csv_neutralizes_a_formula_detail_end_to_end() -> None:
    results = [RowResult(MATRIX[0], Status.FAIL, "=cmd|'/c calc'!A1", "\x00=evidence()")]
    rows = list(csv.reader(io.StringIO(render_csv(results), newline="")))
    assert rows[1][6] == "'=cmd|'/c calc'!A1"
    assert rows[1][7] == "'\x00=evidence()"


# --- write_xlsx_status's WIRING, asserted without the library ----------------
#
# The three tests above pin the helper; these pin that ``write_xlsx_status`` actually CALLS it. Without
# them, reverting the call site to ``value=result.detail`` and deleting the ``data_type`` pin is green
# everywhere, because the two end-to-end tests below (and the pre-existing one in
# tests/test_acceptance_framework.py) importorskip a library that is installed on no CI leg. openpyxl
# is stubbed rather than skipped: the fake records what the writer asked to store.


class _FakeCell:
    """The two attributes ``write_xlsx_status`` touches on a cell it wrote."""

    def __init__(self, row: int, column: int, value: object) -> None:
        self.row = row
        self.column = column
        self.value = value
        self.data_type: str | None = None


class _FakeSheet:
    def __init__(self, rows: list[list[object]]) -> None:
        self._grid = [
            [_FakeCell(r, c, v) for c, v in enumerate(values, start=1)]
            for r, values in enumerate(rows, start=1)
        ]
        self.written: dict[tuple[int, int], _FakeCell] = {}

    @property
    def max_row(self) -> int:
        return len(self._grid)

    def iter_rows(
        self, min_row: int = 1, max_row: int | None = None
    ) -> list[tuple[_FakeCell, ...]]:
        end = self.max_row if max_row is None else min(max_row, self.max_row)
        return [tuple(row) for row in self._grid[min_row - 1 : end]]

    def cell(self, row: int, column: int, value: object = None) -> _FakeCell:
        written = _FakeCell(row, column, value)
        self.written[(row, column)] = written
        return written


class _FakeWorkbook:
    def __init__(self, rows: list[list[object]]) -> None:
        self.active = _FakeSheet(rows)
        self.saved = False

    def save(self, path: Path) -> None:
        self.saved = True


def _stub_openpyxl(monkeypatch: pytest.MonkeyPatch) -> _FakeWorkbook:
    """Install a fake ``openpyxl`` whose ``load_workbook`` returns a recording workbook."""
    workbook = _FakeWorkbook([["ID", "Status", "Result / evidence"], [MATRIX[0].id, "", ""]])
    module = types.ModuleType("openpyxl")
    module.load_workbook = lambda _path: workbook  # type: ignore[attr-defined]
    monkeypatch.setitem(sys.modules, "openpyxl", module)
    return workbook


def test_xlsx_write_back_routes_the_detail_through_the_neutralizer(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = _stub_openpyxl(monkeypatch)
    results = [RowResult(MATRIX[0], Status.FAIL, "=cmd|'/c calc'!A1")]

    assert write_xlsx_status(results, tmp_path / "matrix.xlsx") == 1

    written = workbook.active.written[(2, 3)]
    assert written.value == "'=cmd|'/c calc'!A1", "the detail was stored unneutralized"
    assert written.data_type == "s", "the cell was not pinned to the string type"
    assert workbook.saved


def test_xlsx_write_back_drops_a_control_character_the_library_refuses(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = _stub_openpyxl(monkeypatch)
    results = [RowResult(MATRIX[0], Status.ERROR, "\x00=evil()")]

    assert write_xlsx_status(results, tmp_path / "matrix.xlsx") == 1

    written = workbook.active.written[(2, 3)]
    assert written.value == "'=evil()"  # NUL dropped, and the exposed '=' neutralized
    assert "\x00" not in str(written.value)
    assert written.data_type == "s"


def test_xlsx_write_back_never_stores_a_live_formula(tmp_path: Path) -> None:
    # End-to-end confirmation against the real library where it is installed (the WIN2025 acceptance
    # box only — see above). The stub tests are what hold the wiring in CI; this proves the stub's
    # expectations match openpyxl's actual behaviour.
    openpyxl = pytest.importorskip("openpyxl")
    xlsx = tmp_path / "matrix.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Status", "Result"])
    ws.append([MATRIX[0].id, "", ""])
    wb.save(xlsx)

    results = [RowResult(MATRIX[0], Status.FAIL, "=cmd|'/c calc'!A1")]
    assert write_xlsx_status(results, xlsx) == 1

    cell = openpyxl.load_workbook(xlsx).active.cell(row=2, column=3)
    assert cell.data_type == "s", "detail was stored as a live formula"
    assert "cmd" in str(cell.value)  # anti-vacuity: the detail really was written


def test_xlsx_write_back_survives_a_control_character_detail(tmp_path: Path) -> None:
    openpyxl = pytest.importorskip("openpyxl")
    xlsx = tmp_path / "matrix.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Status", "Result"])
    ws.append([MATRIX[0].id, "", ""])
    wb.save(xlsx)

    # Pre-fix this raised openpyxl.utils.exceptions.IllegalCharacterError out of the CLI.
    results = [RowResult(MATRIX[0], Status.ERROR, "\x00=evil()")]
    assert write_xlsx_status(results, xlsx) == 1
    cell = openpyxl.load_workbook(xlsx).active.cell(row=2, column=3)
    assert cell.data_type == "s"
    assert "\x00" not in str(cell.value)


def test_repo_root_scans_exclude_nested_worktrees() -> None:
    """Any gate that walks from the REPO ROOT must skip `.claude/`, or it scans a second checkout.

    docs/WORKTREES.md puts sibling worktrees at `.claude/worktrees/<name>/` — inside this tree. A
    repo-root `rglob` therefore sees another session's entire working copy, at a different commit.
    Both directions are wrong: it can fail on a file this checkout does not contain (which is how this
    was found — 5,712 stray .py files), and a gate that searches for a REQUIRED registration could find
    it over there and pass while this checkout lacks it. The second is the dangerous one, and it is
    invisible in CI because CI has no nested worktrees.

    Scoped deliberately to repo-root walkers. Narrow roots (`messagefoundry/`, `samples/messages/`)
    cannot reach `.claude/` and are left alone rather than blanket-patched. A first pass at sizing this
    counted 13 files "using rglob"; only ONE walks from the repo root, which is the question that
    matters.

    SCOPE OF THE TWO ASSERTIONS, stated because they differ. The `_SKIP_DIRS` membership check is the
    always-on guard and is what mutation-testing confirms. The `leaked` check below only has teeth in a
    tree that actually HAS a nested worktree — in a sibling worktree (or in CI) there is nothing under
    `.claude/` to find, so it passes trivially. That is acceptable for a belt-and-braces assertion, but
    it is not evidence on its own, and it should not be read as one.
    """
    assert ".claude" in _SKIP_DIRS, (
        "a repo-root scan that does not skip .claude/ will walk a sibling worktree's full checkout"
    )
    # And prove the exclusion actually bites: no scanned path may sit under .claude/.
    scanned = [p for p in _REPO.rglob("*.py") if not _SKIP_DIRS & set(p.parts)]
    leaked = [str(p.relative_to(_REPO)) for p in scanned if ".claude" in p.parts]
    assert not leaked, f"nested-worktree files reached the scan: {leaked[:5]}"
    # Non-vacuity: the scan must actually be finding this repo's own files.
    assert len(scanned) > 500, (
        f"only {len(scanned)} files scanned — the walk collapsed, so a pass proves nothing"
    )
